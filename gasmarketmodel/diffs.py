'''
Module to calculate differences between scenarios
'''

import pandas as pd
import os
from params import OUTPUT_FOLDER, output_dict, OUTPUT_TEMPLATE_FILE
import shutil
import openpyxl

# Find all scenarios
scenarios = [scenario for scenario in os.listdir(OUTPUT_FOLDER / "scenarios") if os.path.isdir(os.path.join(OUTPUT_FOLDER / "scenarios", scenario)) and os.path.isfile(OUTPUT_FOLDER / "scenarios" /  scenario / f"output_{scenario}.xlsx")]

# Initialise empty dictionary
scenario_dict = {}
scenario_price_dict = {}

# Iterate over all scenarios
for scenario_index, scenario in enumerate(scenarios):
    # Read in data for scenario
    scenario_dict[scenario] = {}
    
    # Read in data from each sheet
    for output_metric, output_params in output_dict.items():
        scenario_dict[scenario][output_metric] = pd.read_excel(
            io = OUTPUT_FOLDER / "scenarios" /  scenario / f"output_{scenario}.xlsx",
            sheet_name = output_metric,
            skiprows = 4 + output_params[1],
            index_col = output_params[2]
        )
        
    # Read in total prices
    scenario_price_dict[scenario] = {}
    wb = openpyxl.load_workbook(OUTPUT_FOLDER / "scenarios" /  scenario / f"output_{scenario}.xlsx")
    ws = wb["Charting"]
    for col in range(4, 16):
        scenario_price_dict[scenario][ws.cell(row = 1, column = col).value] = scenario_dict[scenario]["Cost"].iloc[-1, col-2] / scenario_dict[scenario]["Demand"].iloc[-1, col-2] / int(ws.cell(row = 2, column = col).value) * 1000
    scenario_price_dict[scenario][ws.cell(row = 1, column = 17).value] = scenario_dict[scenario]["Cost"].iloc[-1, 17] / scenario_dict[scenario]["Demand"].iloc[-1, 17] / int(ws.cell(row = 2, column = 17).value) * 1000
    scenario_price_dict[scenario][ws.cell(row = 1, column = 18).value] = scenario_dict[scenario]["Cost"].iloc[-1, 18] / scenario_dict[scenario]["Demand"].iloc[-1, 18] / int(ws.cell(row = 2, column = 18).value) * 1000
    scenario_price_dict[scenario][ws.cell(row = 1, column = 20).value] = scenario_dict[scenario]["Cost"].iloc[-1, 19] / scenario_dict[scenario]["Demand"].iloc[-1, 19] / int(ws.cell(row = 2, column = 20).value) * 1000
    
# Subtract scenarios
#TODO - make list
base_scenario = "No NS2"
alt_scenario = "Master"

# Initialise empty difference df dict
diff_df_dict = {}
for output_metric in output_dict.keys():
    diff_df_dict[output_metric] = scenario_dict[alt_scenario][output_metric].subtract(scenario_dict[base_scenario][output_metric])

# Output file name
try:
    os.mkdir(OUTPUT_FOLDER / "deltas" / f"Delta_{alt_scenario}_{base_scenario}")
except FileExistsError:
    pass
output_file = OUTPUT_FOLDER / "deltas" / f"Delta_{alt_scenario}_{base_scenario}" / f"output_Delta_{alt_scenario}_{base_scenario}.xlsx"
# Copy template
shutil.copyfile(OUTPUT_TEMPLATE_FILE, output_file)
# Excel Writer
wb = openpyxl.load_workbook(output_file)
for output_key, output_data in diff_df_dict.items():
    ws = wb[output_key]
    index_cols = output_data.index.tolist()
    value_cols = output_data.values.tolist()
    index_row = output_data.columns.tolist()
    for row_index, row in enumerate(value_cols):
        # Turn single-column index into list to enable for loop below
        if not isinstance(index_cols[row_index], tuple):
            index_cols[row_index] = [index_cols[row_index]]
        for index_col_index, index_col in enumerate(index_cols[row_index]):
            ws.cell(row = row_index + 6 + output_dict[output_key][1], column = index_col_index + 1, value = index_col)
            # On a multi-index, delete additional indices
            if row_index == len(value_cols) - 1 and output_dict[output_key][0] and index_col_index != 0:
                ws.cell(row = row_index + 6 + output_dict[output_key][1], column = index_col_index + 1, value = "")
            for col_index, col in enumerate(row):
                ws.cell(row = row_index + 6 + output_dict[output_key][1], column = col_index + 1 + len(index_cols[row_index]), value = col)
                ws.cell(row = 5 + output_dict[output_key][1], column = col_index + 1 + len(index_cols[row_index]), value = index_row[col_index])
                # Make total row bold
                if row_index == len(value_cols) - 1 and output_dict[output_key][0]:
                    ws.cell(row = row_index + 6 + output_dict[output_key][1], column = col_index + 1 + len(index_cols[row_index])).font = openpyxl.styles.Font(bold = True)
                
# Scenario name
ws = wb["Scenario"]
ws.cell(1, 3, value = f"Delta_{alt_scenario}_{base_scenario}")

# Total Price delta
ws = wb["Charting"]
for col in range(4, 21):
    if isinstance(ws.cell(row = 1, column = col).value, str):
        ws.cell(row = 10, column = col, value = scenario_price_dict[alt_scenario][ws.cell(row = 1, column = col).value] - scenario_price_dict[base_scenario][ws.cell(row = 1, column = col).value])
        
# Save
wb.save(output_file)     