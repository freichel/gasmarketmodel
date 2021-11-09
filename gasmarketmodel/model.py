'''
Main module to create scenario outputs
'''

from params import SCENARIO_FOLDER, OUTPUT_FOLDER, OUTPUT_TEMPLATE_FILE, seasons_df
import excelimport as ei
import os
import pandas as pd
import pulp
import numpy as np
import shutil
import openpyxl

# Find all scenario files
print("Finding scenarios...")
scenario_file_list = []
for file in os.listdir(SCENARIO_FOLDER):
    if file.endswith(".xlsm") and not file.startswith("~"):
        scenario_file_list.append(SCENARIO_FOLDER / file)
print("...scenarios loaded")
print("")

'''
Scenario-unspecific imports - use first scenario file as reference
'''
print("Importing general data...")
# Cycles 
cycles_days_dict = ei.import_excel_generic_df(SCENARIO_FOLDER / scenario_file_list[0], "Cycles", 4, 0).fillna(0).iloc[0].to_dict()
# Unit conversions
unit_conversions_df = ei.import_excel_generic_df(SCENARIO_FOLDER / scenario_file_list[0], "Unit Conversions", 6, 0)
# Forex conversions
forex_conversions_df = ei.import_excel_generic_df(SCENARIO_FOLDER / scenario_file_list[0], "Forex Conversions", 4, 0)
# Region defs
regions_df = ei.import_excel_generic_df(SCENARIO_FOLDER / scenario_file_list[0], "Regions > Index", 4, 1)
regions_df["Master"].fillna(regions_df["Region"], inplace = True)
# Piped importers defs
piped_importers_df = ei.import_excel_generic_df(SCENARIO_FOLDER / scenario_file_list[0], "Piped Importers > Index", 4, 1)
# Piped importers connections
piped_importers_connections_df = ei.import_excel_generic_df(SCENARIO_FOLDER / scenario_file_list[0], "Piped Importers > Connections", 4, 8).merge(
    regions_df[["Region", "Master"]],
    on = "Region"
)
piped_importers_connections_df["Region"] = piped_importers_connections_df["Master"]
piped_importers_connections_df.drop(columns = ["Master"], inplace = True)
# Piped exporters defs
piped_exporters_df = ei.import_excel_generic_df(SCENARIO_FOLDER / scenario_file_list[0], "Piped Exporters > Index", 4, 1)
# Piped exporters connections
piped_exporters_connections_df = ei.import_excel_generic_df(SCENARIO_FOLDER / scenario_file_list[0], "Piped Exporters > Connections", 4, 8).merge(
    regions_df[["Region", "Master"]],
    on = "Region"
)
piped_exporters_connections_df["Region"] = piped_exporters_connections_df["Master"]
piped_exporters_connections_df.drop(columns = ["Master"], inplace = True)
# LNG importers defs
lng_importers_df = ei.import_excel_generic_df(SCENARIO_FOLDER / scenario_file_list[0], "LNG Importers > Index", 4, 1)
# LNG importers connections
lng_importers_connections_df = ei.import_excel_generic_df(SCENARIO_FOLDER / scenario_file_list[0], "LNG Importers > Connections", 4, 8).merge(
    regions_df[["Region", "Master"]],
    on = "Region"
)
lng_importers_connections_df["Region"] = lng_importers_connections_df["Master"]
lng_importers_connections_df.drop(columns = ["Master"], inplace = True)
# Region connections defs
connections_df = ei.import_excel_generic_df(SCENARIO_FOLDER / scenario_file_list[0], "Connections > Index", 4, 8).merge(
    regions_df[["Region", "Master"]],
    left_on = "Origin",
    right_on = "Region",
    suffixes = ("", "_o")
).merge(
    regions_df[["Region", "Master"]],
    left_on = "Destination",
    right_on = "Region",
    suffixes = ("", "_d")
)
connections_df["Origin"] = connections_df["Master"]
connections_df["Destination"] = connections_df["Master_d"]
connections_df.drop(columns = ["Master", "Master_d", "Region", "Region_d"], inplace = True)

print("...general data imported")
print("")

print("Scenario modelling")
# Loop over scenario files
for scenario_file in scenario_file_list:
    '''
    Scenario specific imports from Excel scenario file
    '''
    # Scenario name
    scenario = ei.import_excel_cell(scenario_file, "Scenario", 4, 2)
    print(f" Modelling scenario {scenario}...")
    
    print("  Importing data...")
    # Region imports
    # Region demand data
    regions_demand_df = ei.import_excel_generic_df(scenario_file, "Regions > Demand", 4, 1).merge(
        regions_df[["Region", "Master"]],
        on = "Region"
    ).groupby("Master").sum().reset_index().rename(columns = {"Master" : "Region"}).set_index("Region")
    # Region production data
    regions_production_df = ei.import_excel_generic_df(scenario_file, "Regions > Production", 4, 1).merge(
        regions_df[["Region", "Master"]],
        on = "Region"
    ).groupby("Master").sum().reset_index().rename(columns = {"Master" : "Region"}).set_index("Region")
    
    # Piped importers
    # Piped importers production
    piped_importers_production_df = ei.import_excel_generic_df(scenario_file, "Piped Importers > Production", 4, 1)
    # Piped importers connections data
    piped_importers_connections_data_df = ei.import_excel_connections_data_df(scenario_file, "Piped Importers > Data", 4, [0,1], piped_importers_connections_df)
    
    # Piped exporters
    # Piped exporters demand
    piped_exporters_demand_df = ei.import_excel_generic_df(scenario_file, "Piped Exporters > Demand", 4, 1)
    
    # LNG importers
    # LNG importers connections data
    lng_importers_connections_data_df = ei.import_excel_connections_data_df(scenario_file, "LNG Importers > Data", 4, [0,1])
    
    # Region connections
    # Region connections data
    connections_data_df = ei.import_excel_connections_data_df(scenario_file, "Connections > Data", 4, [0,1])
    
    # Storage
    # Storage defs
    storage_df = ei.import_excel_generic_df(scenario_file, "Storage > Index", 4, 0).merge(
        regions_df[["Region", "Master"]],
        right_on = "Region",
        left_index = True
    ).fillna(-1).groupby("Master", dropna = False).sum().reset_index().rename(columns = {"Master" : "Region"}).set_index("Region")
    # Storage target balance
    storage_target_balance_df = ei.import_excel_generic_df(scenario_file, "Storage > Target Balance", 4, 0)
    # Storage volumes
    storage_volumes_df = pd.DataFrame(index = storage_df.index.values, columns = cycles_days_dict.keys())
    # Obtain individual target balance for each storage
    for storage_index, storage_data in storage_df.iterrows():
        for cycle, cycle_days in cycles_days_dict.items():
            # Column index
            col_index = list(storage_df.columns.values).index(cycle)
            # Above we replaced NaN with negative values (which cannot exist).
            # We can use this to now generate the actual targets.
            if storage_df.loc[storage_index][cycle] < 0:
                storage_df.at[storage_index, cycle] = round(storage_df.loc[storage_index]["WGV"] * storage_target_balance_df.loc["Marktgebiet"][cycle], 2)
            # Daily injection or withdrawal volume
            storage_volumes_df.at[storage_index, cycle] = round((storage_df.loc[storage_index, storage_df.columns.values[col_index - 1]] - storage_df.loc[storage_index, cycle]) / cycle_days, 2)
    storage_volumes_df.fillna(value = 0, inplace = True)
    
    # LNG
    # LNG supply curve
    lng_supply_curve_df = ei.import_excel_generic_df(scenario_file, "LNG > Price Curve", 4, 0)
    # Other LNG demand
    lng_other_demand_dict = ei.import_excel_generic_df(scenario_file, "LNG > Other Demand", 4, 0).fillna(0).iloc[0].to_dict()
    print("  ...data imported")
    
    '''
    Define output dataframes
    '''
    output_demand_df = regions_demand_df.drop(columns = ["Source/Comments", "DemandID"], errors = "ignore")
    output_production_df = regions_production_df.drop(columns = ["Source/Comments", "ProductionID"], errors = "ignore")
    output_piped_importers_df = pd.DataFrame()
    output_piped_exporters_df = pd.DataFrame()
    output_lng_importers_df = pd.DataFrame()
    output_storage_df = storage_df.drop(storage_df.columns[1], axis = 1).drop(columns = ["Comments", "StorageID"], errors = "ignore")
    output_storage_volumes_df = storage_volumes_df.copy()
    output_connections_df = pd.DataFrame()
    output_prices_df = pd.DataFrame()
    output_supply_mix_df = pd.DataFrame(
        0,
        pd.MultiIndex.from_product(
        [
            output_demand_df.index.values.tolist(),
            ["Production", "Piped Imports", "Piped Exports", "LNG Imports", "Transit Imports", "Transit Exports", "Storage"]
        ],
        names = ["Region", "Type"]
        ),
        []
    )
    
    
    '''
    Analysis for each cycle
    '''
    # Counter variable
    cycle_index = 0
    print("  Starting analysis...")
    for cycle, cycle_days in cycles_days_dict.items():
        print(f"   Cycle {cycle} ({cycle_index + 1} of {len(cycles_days_dict)})...")
        '''
        Model preparation
        '''
        # Set up model variables
        # Piped importers
        piped_importers_connections_dict = {}
        for piped_importers_connection_index, piped_importers_connection in piped_importers_connections_df.iterrows():
            piped_importers_connections_dict[piped_importers_connection_index] = pulp.LpVariable(
                piped_importers_connection["Name"],
                lowBound = 0,
                upBound = (
                    piped_importers_connections_data_df.loc[
                        piped_importers_connection["Name"],
                        "Capacity - Max"]
                    [cycle]
                    - piped_importers_connections_data_df.loc[
                        piped_importers_connection["Name"],
                        "Capacity - Min"]
                    [cycle]
                )
            )
        # Piped exporters
        piped_exporters_connections_dict = {}
        for piped_exporters_connection_index, piped_exporters_connection in piped_exporters_connections_df.iterrows():
            piped_exporters_connections_dict[piped_exporters_connection_index] = pulp.LpVariable(
                piped_exporters_connection["Name"],
                lowBound = 0,
            )
        # LNG importers
        lng_importers_connections_dict = {}
        for lng_importers_connection_index, lng_importers_connection in lng_importers_connections_df.iterrows():
            lng_importers_connections_dict[lng_importers_connection_index] = pulp.LpVariable(
                lng_importers_connection["Name"],
                lowBound = 0,
                upBound = (
                    lng_importers_connections_data_df.loc[
                        lng_importers_connection["Name"],
                        "Capacity - Max"]
                    [cycle]
                    - lng_importers_connections_data_df.loc[
                        lng_importers_connection["Name"],
                        "Capacity - Min"]
                    [cycle]
                )
            )
        # Region connections
        connections_dict = {}
        for connection_index, connection in connections_df.iterrows():
            connections_dict[connection_index] = pulp.LpVariable(
                connection["Name"],
                lowBound = 0,
                upBound = (
                    connections_data_df.loc[
                        connection["Name"],
                        "Capacity - Max"]
                    [cycle]
                    - connections_data_df.loc[
                        connection["Name"],
                        "Capacity - Min"]
                    [cycle]
                )
            )
        
        # Regional overviews
        # Dictionary ountry: [factor, connection, fixed tariff, variable tariff, lng source switch, name, source]
        regions_supply_dict = {}
        # Dictionary country: demand
        regions_demand_dict = {}
        for region_index, region in regions_df[regions_df["Region"] == regions_df["Master"]].set_index("Region").iterrows():
            # Demand
            region_demand = regions_demand_df.loc[region_index][cycle]
            regions_demand_dict[region_index] = region_demand
            
            # Production
            if region_index in regions_production_df.index:
                region_production = regions_production_df.loc[region_index][cycle]
            else:
                region_production = 0
            regions_supply_dict[region_index] = [[1, region_production, 0, 0, 0, f"Production {region_index}", ""]]
            
            # Piped imports
            piped_importer_list = piped_importers_connections_df[piped_importers_connections_df["Region"] == region_index]["Name"].to_dict()
            for piped_importer_id, piped_importer in piped_importer_list.items():
                # Min flow
                regions_supply_dict[region_index].append([1, piped_importers_connections_data_df.loc[piped_importer, "Capacity - Min"][cycle], 0, 0, 0, f"{piped_importer} - Min", ""])
                # Variable flow
                regions_supply_dict[region_index].append(
                    [
                        1,
                        piped_importers_connections_dict[piped_importer_id],
                        piped_importers_connections_data_df.loc[piped_importer, "Cost to Market - fixed"][cycle],
                        piped_importers_connections_data_df.loc[piped_importer, "Cost to Market - variable"][cycle] / unit_conversions_df.loc["GWh"]["MWh"],
                        0,
                        piped_importer,
                        "Pipe"
                    ]
                )
                
            # Piped exports
            piped_exporter_list = piped_exporters_connections_df[piped_exporters_connections_df["Region"] == region_index]["Name"].to_dict()
            for piped_exporter_id, piped_exporter in piped_exporter_list.items():
                # Variable flow
                regions_supply_dict[region_index].append([-1, piped_exporters_connections_dict[piped_exporter_id], 0, 0, 0, piped_exporter, "Pipe"])
            
            # Storage
            if region_index in storage_volumes_df.index:
                region_storage = storage_volumes_df.loc[region_index][cycle]
                if region_storage != 0:
                    direction = region_storage/abs(region_storage)
                else:
                    direction = 1
                regions_supply_dict[region_index].append([direction, abs(region_storage), 0, 0, 0, f"Storage {region_index}", ""])
                
            # LNG imports
            lng_importer_list = lng_importers_connections_df[lng_importers_connections_df["Region"] == region_index]["Name"].to_dict()
            for lng_importer_id, lng_importer in lng_importer_list.items():
                # Min flow
                regions_supply_dict[region_index].append([1, lng_importers_connections_data_df.loc[lng_importer, "Capacity - Min"][cycle], 0, 0, 0, f"{lng_importer} - Min", ""])
                # Variable flow
                regions_supply_dict[region_index].append(
                    [
                        1,
                        lng_importers_connections_dict[lng_importer_id],
                        lng_importers_connections_data_df.loc[lng_importer, "Cost to Market - fixed"][cycle],
                        lng_importers_connections_data_df.loc[lng_importer, "Cost to Market - variable"][cycle] / unit_conversions_df.loc["GWh"]["MWh"],
                        1,
                        lng_importer,
                        "LNG"
                    ]
                )
            
            # Region connections (imports)
            import_list = connections_df[connections_df["Destination"] == region_index]["Name"].to_dict()
            for importer_id, importer in import_list.items():
                # Min flow
                regions_supply_dict[region_index].append([1, connections_data_df.loc[importer, "Capacity - Min"][cycle], 0, 0, 0, f"{importer} - Min", ""])
                # Variable flow
                regions_supply_dict[region_index].append(
                    [
                        1,
                        connections_dict[importer_id],
                        connections_data_df.loc[importer, "Tariff - fixed"][cycle],
                        connections_data_df.loc[importer, "Tariff - variable"][cycle] / unit_conversions_df.loc["GWh"]["MWh"],
                        0,
                        importer,
                        connections_df[connections_df["Name"] == importer]["Origin"].item()
                    ]
                )
            
            # Region connections (exports)
            export_list = connections_df[connections_df["Origin"] == region_index]["Name"].to_dict()
            for exporter_id, exporter in export_list.items():
                # Min flow
                regions_supply_dict[region_index].append([-1, connections_data_df.loc[exporter, "Capacity - Min"][cycle], 0, 0, 0, f"{exporter} - Min", ""])
                # Variable flow
                regions_supply_dict[region_index].append([-1, connections_dict[exporter_id], 0, -connections_data_df.loc[exporter, "Tariff - variable"][cycle] / unit_conversions_df.loc["GWh"]["MWh"], 0, exporter, connections_df[connections_df["Name"] == exporter]["Destination"].item()])


        '''
        Model
        '''
        # Initialise model
        cost_total = pulp.LpProblem("Cost", pulp.LpMinimize)
        
        # Volume balancing
        # Supply/demand matching
        for region_key, region_data in regions_supply_dict.items():
            # Aggregate all supply sources (minus exports) for region
            region_supply = pulp.lpSum([row[0] * row[1] for row in region_data])
            # Add constraints so that supply will exactly match demand
            cost_total += (region_supply <= regions_demand_dict[region_key])
            cost_total += (region_supply >= regions_demand_dict[region_key])
        # Piped imports
        for piped_importer_index, piped_importer in piped_importers_df.iterrows():
            # Aggregate all piped import connections
            importer_supply = pulp.lpSum(
                [piped_importers_connections_dict[piped_importers_connection] for piped_importers_connection in piped_importers_connections_df[piped_importers_connections_df["Importer Index"] == piped_importer_index].index.tolist()]
            )
            # Add constraints so that the sum of flows will not exceed production
            cost_total += (
                importer_supply
                <= (
                    piped_importers_production_df.loc[piped_importer_index][cycle]
                    - piped_importers_connections_data_df[piped_importers_connections_data_df["Importer"] == piped_importer["Importer"]].xs(
                        "Capacity - Min",
                        level = 1,
                        drop_level = False
                    )
                    [cycle].sum())
            )
        # Piped exports
        for piped_exporter_index, piped_exporter in piped_exporters_df.iterrows():
            # Aggregate all piped export connections
            exporter_demand = pulp.lpSum(
                [piped_exporters_connections_dict[piped_exporters_connection] for piped_exporters_connection in piped_exporters_connections_df[piped_exporters_connections_df["Exporter Index"] == piped_exporter_index].index.tolist()]
            )
            # Add constraints so that the sum of flows will match demand
            cost_total += (exporter_demand <= piped_exporters_demand_df.loc[piped_exporter_index][cycle])
            cost_total += exporter_demand >= (piped_exporters_demand_df.loc[piped_exporter_index][cycle])
        # LNG
        # LNG volume
        lng_total = (
            # Demand outside markets
            lng_other_demand_dict[cycle] * unit_conversions_df.loc["Mt LNG"]["GWh"] / cycle_days
            # Demand
            + regions_demand_df[cycle].sum()
            # Production
            - regions_production_df[cycle].sum()
            # Storage
            - storage_volumes_df[cycle].sum()
            # Piped Imports - min of available production and max capacity
            - min(
                piped_importers_production_df[cycle].sum(),
                piped_importers_connections_data_df.xs("Capacity - Max", level = 1, drop_level = False)[cycle].sum()
            )
        ) * unit_conversions_df.loc["GWh"]["Mt LNG"] * cycle_days
        # LNG price curve
        lng_pricelist = lng_supply_curve_df[cycle].to_frame().reset_index().values
        # LNG price at given point
        lng_price = round(lng_pricelist[np.argmin(np.abs(lng_pricelist[0:, 0] - lng_total)), 1] / forex_conversions_df.loc["USD"][cycle] * unit_conversions_df.loc["GWh"]["MMBTU"], 2)
        
        # Cost calculations
        regions_cost_dict = {}
        for region_key, region_data in regions_supply_dict.items():
            # Define binary flags for fixed tariffs
            region_binaries = []
            for row in region_data:
                # Only positive variable flows
                if row[0] == 1 and isinstance(row[1], pulp.pulp.LpVariable):
                    region_binaries.append(pulp.LpVariable(f"Fixed_Cost_Flag_{row[5]}", cat = "Binary"))
                    cost_total += (region_binaries[-1] >= row[1] * 0.001)
                    cost_total += (region_binaries[-1] <= row[1] * 1e10)
                else:
                    region_binaries.append(0)
            # Multiply attracted supplies (1, not -1) by their respective tariffs
            # 0 is factor
            # 1 is volume
            # 2 is fixed tariff
            # 3 is variable tariff
            # 4 is lng switch
            region_cost = pulp.lpSum([row[0] * row[1] * (max(row[3], 0) + row[4] * lng_price) + region_binaries[row_index] * row[2] for row_index, row in enumerate(region_data) if row[0] == 1])
            # Append to dictionary
            regions_cost_dict[region_key] = region_cost
        # Add up total cost and add as objective    
        cost_total += pulp.lpSum([region_data for region_key, region_data in regions_cost_dict.items()])
        
        # Solve
        res = cost_total.solve(pulp.PULP_CBC_CMD(msg = 0))
        assert res == pulp.LpStatusOptimal
        print("    Solution for cycle found!")

        
        '''
        Calculation outputs
        '''
        # Store value for each model variable in dict
        model_vars_dict = {}
        for model_var in cost_total.variables():
            model_vars_dict[model_var.name] = model_var.value()
        
        # Prices
        # Price sources
        regions_sources_dict = {}
        regions_prices_dict = {}
        for region_key, region_data in regions_supply_dict.items():
            regions_sources_dict[region_key] = []
            # Check if any flows are imports. If so, we can discard exports (see below)
            positive_flag = False
            for region_vars in region_data:
                # Only variable flows
                if isinstance(region_vars[1], pulp.pulp.LpVariable):
                    # Only non-zero flows
                    if model_vars_dict[region_vars[1].name] != 0 and region_vars[3] != 0 and region_vars[6] != "Pipe":
                        # Set flag if this is a positive flow
                        if region_vars[3] >= 0:
                            positive_flag = True
                        # Source, cost per MWh imported (only variable)
                        regions_sources_dict[region_key].append(
                            [
                                region_vars[6],
                                region_vars[3] * unit_conversions_df.loc["GWh"]["MWh"]
                            ]
                        )
            # We actually only need to keep exports (negative tariffs) if no positive ones exist (see above)
            if positive_flag == True:
                temp_list = regions_sources_dict[region_key]
                regions_sources_dict[region_key] = [temp_row for temp_row in temp_list if temp_row[1] >= 0]
            regions_prices_dict[region_key] = None
        # Find prices
        # The algorithm always finds the next possible market price (i.e. one with no ambiguity)
        # The region is then removed from the dict
        # Continuous iteration until all solutions are found
        # Start with LNG
        regions_prices_dict["LNG"] = lng_price * unit_conversions_df.loc["GWh"]["MWh"]
        # Keep looping until all values have been found
        stuck = 0
        while None in regions_prices_dict.values():
            # Take a temporary copy to check that there has been progress in solving this
            temp_dict = regions_sources_dict.copy()
            # Iterate over all regions
            for region_key, region_prices in regions_sources_dict.items():
                for price_index, region_price in enumerate(region_prices):
                    # Check if market price has already been found
                    if isinstance(region_price, list):
                        if regions_prices_dict[region_price[0]] != None:
                            regions_sources_dict[region_key][price_index] = regions_prices_dict[region_price[0]] + region_price[1]
                # Now check if all values have been found, and if so save it. Then delete entry from dict and restart
                if all(isinstance(region_price, (int, float)) for region_price in region_prices) and len(region_prices) > 0:
                    regions_prices_dict[region_key] = max(region_prices)
                    del regions_sources_dict[region_key]
                    break
            # If no progress has been made
            if regions_sources_dict == temp_dict:
                if stuck > 0:
                    # Select a random as-yet unsolved region from the dictionary
                    random_region = list(regions_sources_dict)[stuck - 1]
                    # Make sure we filter out the neighboring region it's stuck on
                    if len(regions_sources_dict[random_region]) > 0:
                        stuck_on = regions_sources_dict[random_region][0][0]
                    else:
                        stuck_on = ""
                    # Available connections
                    available_connections_df = connections_df[(connections_df["Origin"] == random_region) & (connections_df["Destination"] != stuck_on)].merge(
                        connections_data_df.xs("Tariff - variable", level = 1, drop_level = False).reset_index(),
                        left_on = "Name",
                        right_on = "Transit Connection"
                    )[["Destination", cycle]].sort_values(by = [cycle])
                    # Iterate over available destinations
                    for _, connection in available_connections_df.iterrows():
                        # Check if this one has already been solved
                        if regions_prices_dict.get(connection["Destination"], None) is not None:
                            regions_sources_dict[random_region].append(regions_prices_dict[connection["Destination"]] - connection[cycle])
                            stuck = 0
                            continue
                stuck += 1
                # Iterate over remaining regions
                for region_key, region_prices in regions_sources_dict.items():
                    # Find first instance of an unresolved price
                    if len(region_prices) > 1:
                        for region_price_index, region_price in enumerate(region_prices):
                            if isinstance(region_price, list):
                                # Delete it
                                del regions_sources_dict[region_key][region_price_index]
                                stuck = False
                                break
                            else:
                                continue
                    else:
                        continue
                    break
        
        '''
        Dataframe outputs
        '''
        # Piped imports
        for piped_importer_id, piped_importer in piped_importers_connections_dict.items():
            piped_importer_name = piped_importers_connections_df.loc[piped_importer_id]["Name"]
            # Only in first cycle
            if cycle_index == 0:
                output_piped_importers_df.at[
                    piped_importer_name,
                    "Importer"
                ] = piped_importers_connections_df.loc[piped_importer_id]["Importer"]
                output_piped_importers_df.at[
                    piped_importer_name,
                    "Region"
                ] = piped_importers_connections_df.loc[piped_importer_id]["Region"]
            output_piped_importers_df.at[
                piped_importer_name,
                cycle
            ] = piped_importer.value() + piped_importers_connections_data_df.loc[piped_importer_name, "Capacity - Min"][cycle]
        
        # Piped exports
        for piped_exporter_id, piped_exporter in piped_exporters_connections_dict.items():
            piped_exporter_name = piped_exporters_connections_df.loc[piped_exporter_id]["Name"]
            # Only in first cycle
            if cycle_index == 0:
                output_piped_exporters_df.at[
                    piped_exporter_name,
                    "Exporter"
                ] = piped_exporters_connections_df.loc[piped_exporter_id]["Exporter"]
                output_piped_exporters_df.at[
                    piped_exporter_name,
                    "Region"
                ] = piped_exporters_connections_df.loc[piped_exporter_id]["Region"]
            output_piped_exporters_df.at[
                piped_exporter_name,
                cycle
            ] = - piped_exporter.value()
        
        # LNG imports
        for lng_importer_id, lng_importer in lng_importers_connections_dict.items():
            lng_importer_name = lng_importers_connections_df.loc[lng_importer_id]["Name"]
            # Only in first cycle
            if cycle_index == 0:
                output_lng_importers_df.at[
                    lng_importer_name,
                    "Terminal"
                ] = lng_importers_connections_df.loc[lng_importer_id]["Terminal"]
                output_lng_importers_df.at[
                    lng_importer_name,
                    "Region"
                ] = lng_importers_connections_df.loc[lng_importer_id]["Region"]
            output_lng_importers_df.at[
                lng_importer_name,
                cycle
            ] = lng_importer.value() + lng_importers_connections_data_df.loc[lng_importer_name, "Capacity - Min"][cycle]
        
        # Region connections
        for connection_id, connection in connections_dict.items():
            connection_name = connections_df.loc[connection_id]["Name"]
            # Only in first cycle
            if cycle_index == 0:
                output_connections_df.at[
                    connection_name,
                    "Origin"
                ] = connections_df.loc[connection_id]["Origin"]
                output_connections_df.at[
                    connection_name,
                    "Destination"
                ] = connections_df.loc[connection_id]["Destination"]
            output_connections_df.at[
                connection_name,
                cycle
            ] = connection.value() + connections_data_df.loc[connection_name, "Capacity - Min"][cycle]
            
        # Prices
        # Only in first cycle
        if cycle_index == 0:
            output_prices_df = output_prices_df.from_dict(regions_prices_dict, orient = "index", columns = [cycle])
        else:
            output_prices_df[cycle] = pd.Series(regions_prices_dict)
        
        # Regional supply mix
        output_supply_mix_dict = {}
        output_supply_mix_df[cycle] = 0

        for region_key, region_data in regions_demand_dict.items():
            demand = output_demand_df.loc[region_key][cycle]
            exports = -output_connections_df[output_connections_df["Origin"] == region_key][cycle].sum()
            output_supply_mix_df.loc[(region_key, "Demand"), cycle] = demand
            output_supply_mix_df.loc[(region_key, "Transit Exports"), cycle] = exports
            if region_key in regions_production_df.index:
                output_supply_mix_df.loc[(region_key, "Production"), cycle] = output_production_df.loc[region_key][cycle]
            else:
                output_supply_mix_df.loc[(region_key, "Production"), cycle] = 0
            output_supply_mix_df.loc[(region_key, "Piped Imports"), cycle] = output_piped_importers_df[output_piped_importers_df["Region"] == region_key][cycle].sum()
            output_supply_mix_df.loc[(region_key, "Piped Exports"), cycle] = output_piped_exporters_df[output_piped_exporters_df["Region"] == region_key][cycle].sum()
            output_supply_mix_df.loc[(region_key, "LNG Imports"), cycle] = output_lng_importers_df[output_lng_importers_df["Region"] == region_key][cycle].sum()
            if region_key in output_storage_volumes_df.index:
                output_supply_mix_df.loc[(region_key, "Storage"), cycle] = output_storage_volumes_df.loc[region_key][cycle]
            output_supply_mix_df.loc[(region_key, "Transit Imports"), cycle] = output_connections_df[output_connections_df["Destination"] == region_key][cycle].sum()
        
        print(f"   ...cycle {cycle} ({cycle_index + 1} of {len(cycles_days_dict)}) completed")
        cycle_index += 1
        
    print("  Calculating seasons...")
    # Seasons
    for season in set(seasons_df.index.values):
        season_days = 0
        season_demand = 0
        season_production = 0
        season_piped_importers = 0
        season_piped_exporters = 0
        season_lng_importers = 0
        season_storage_volumes = 0
        season_connections = 0
        season_prices = 0
        season_supply_mix = 0
        for season_index, season_data in seasons_df.loc[season].iterrows():
            # Cycle name
            season_cycle = season_data.values[0]
            # Days
            season_days += cycles_days_dict[season_cycle]
            # Demand
            season_demand += output_demand_df[season_cycle] * cycles_days_dict[season_cycle]
            # Production
            season_production += output_production_df[season_cycle] * cycles_days_dict[season_cycle]
            # Piped importers
            season_piped_importers += output_piped_importers_df[season_cycle] * cycles_days_dict[season_cycle]
            # Piped exporters
            season_piped_exporters += output_piped_exporters_df[season_cycle] * cycles_days_dict[season_cycle]
            # LNG
            season_lng_importers += output_lng_importers_df[season_cycle] * cycles_days_dict[season_cycle]
            # Storage volumes
            season_storage_volumes += output_storage_volumes_df[season_cycle] * cycles_days_dict[season_cycle]
            # Region connections
            season_connections += output_connections_df[season_cycle] * cycles_days_dict[season_cycle]
            # Prices
            season_prices += (output_demand_df[season_cycle] * cycles_days_dict[season_cycle] * output_prices_df[season_cycle]).fillna(output_prices_df[season_cycle] * cycles_days_dict[season_cycle])
            # Supply mix
            season_supply_mix += output_supply_mix_df[season_cycle] * cycles_days_dict[season_cycle]
        # Demand
        output_demand_df[season] = season_demand / season_days
        # Production
        output_production_df[season] = season_production / season_days
        # Piped importers
        output_piped_importers_df[season] = season_piped_importers / season_days
        # Piped exporters
        output_piped_exporters_df[season] = season_piped_exporters / season_days
        # LNG
        output_lng_importers_df[season] = season_lng_importers / season_days
        # Storage fills - last value in season
        output_storage_df[season] = output_storage_df[season_cycle]
        # Storage volumes
        output_storage_volumes_df[season] = season_storage_volumes / season_days
        # Region connections
        output_connections_df[season] = season_connections / season_days
        # Prices
        output_prices_df[season] = (season_prices / season_demand).fillna(season_prices / season_days)
        # Supply mix
        output_supply_mix_df[season] = season_supply_mix / season_days
    print(" ...seasons calculated")
            
    print("  Writing data...")
    '''
    File output
    '''
    # Dictionary with all output dataframes.
    # label: [dataframe, total_row_switch, row offset]
    output_dict = {
        "Demand": [output_demand_df.sort_index(), True, 0],
        "Production": [output_production_df.sort_index(), True, 0],
        "Price": [output_prices_df.sort_index(), False, 0],
        "Storage Volumes": [output_storage_volumes_df.sort_index(), True, 0],
        "Storage Levels": [output_storage_df.sort_index(), True, 1],
        "LNG": [output_lng_importers_df.sort_values(["Terminal", "Region"]), True, 8],
        "Piped Imports": [output_piped_importers_df.sort_values(["Importer", "Region"]), True, 12],
        "Piped Exports": [output_piped_exporters_df.sort_values(["Exporter", "Region"]), True, 12],
        "Connections": [output_connections_df.sort_values(["Origin", "Destination"]), False, 12],
        "Supply Mix": [output_supply_mix_df.sort_index().reset_index().set_index("Region"), False, 12]
    }
    # Output file name
    try:
        os.mkdir(OUTPUT_FOLDER / "scenarios" / scenario)
    except FileExistsError:
        pass
    output_file = OUTPUT_FOLDER / "scenarios" / scenario / "output.xlsx"
    # Copy template
    shutil.copyfile(OUTPUT_TEMPLATE_FILE, output_file)
    # Excel Writer
    wb = openpyxl.load_workbook(output_file)
    for output_key, output_data in output_dict.items():
        ws = wb[output_key]
        # Add total row if required
        if output_data[1]:
            index_col = output_data[0].append(output_data[0].sum(numeric_only = True).rename("Total")).index.tolist()
            value_cols = output_data[0].append(output_data[0].sum(numeric_only = True).rename("Total")).values.tolist()
        else:
            index_col = output_data[0].index.tolist()
            value_cols = output_data[0].values.tolist()
        index_row = output_data[0].columns.tolist()
        for row_index, row in enumerate(value_cols):
            ws.cell(row = row_index + 6 + output_data[2], column = 1, value = index_col[row_index])
            for col_index, col in enumerate(row):
                ws.cell(row = row_index + 6 + output_data[2], column = col_index + 2, value = col, )
                ws.cell(row = 5 + output_data[2], column = col_index + 2, value = index_row[col_index])
                # Make total row bold
                if row_index == len(value_cols) - 1 and output_data[1]:
                    ws.cell(row = row_index + 6 + output_data[2], column = col_index + 2).font = openpyxl.styles.Font(bold = True)
                    
    # Scenario name
    ws = wb["Scenario"]
    ws.cell(1, 3, value = scenario)

    # Save
    wb.save(output_file)     
    print("  ...data written") 
    print(" ...scenario modelled")

print("")
print("Exporting general data...")
# Output file name
output_file = OUTPUT_FOLDER / "params.xlsx"
# Excel Writer
with pd.ExcelWriter(path = output_file, mode = "a", if_sheet_exists = "replace") as writer:
    # Piped importers coordinates
    piped_importers_df[["Importer", "Latitude", "Longitude"]].set_index("Importer").to_excel(writer, sheet_name = "C Piped Importers")
    # Piped importer connections
    piped_importers_connections_df[["Name", "Importer", "Region", "Destination Latitude", "Destination Longitude"]].set_index("Name").to_excel(writer, sheet_name = "C Piped Importers Conn")
    # LNG importer connections
    lng_importers_connections_df[["Name", "Terminal", "Region", "Latitude", "Longitude", "Angle"]].set_index("Name").to_excel(writer, sheet_name = "C LNG Importers Conn")
    # Region connections
    connections_df[["Name", "Origin", "Destination", "Latitude", "Longitude", "Angle"]].set_index("Name").to_excel(writer, sheet_name = "C Connections")
    # Piped importers coordinates
    piped_exporters_df[["Exporter", "Latitude", "Longitude"]].set_index("Exporter").to_excel(writer, sheet_name = "C Piped Exporters")
    # Piped exporter connections
    piped_exporters_connections_df[["Name", "Exporter", "Region", "Origin Latitude", "Origin Longitude"]].set_index("Name").to_excel(writer, sheet_name = "C Piped Exporters Conn")
    # Region mapping
    regions_df[["Master", "Region"]].set_index("Master").to_excel(writer, sheet_name = "Mapping")
    # Cycles
    cycle_list = list(cycles_days_dict.keys())
    cycle_list.extend(set(seasons_df.index.values))
    pd.DataFrame(cycle_list).reset_index().set_index(0).drop(columns = ["index"], errors = "ignore").to_excel(writer, sheet_name = "Cycles")
print("...general data exported")
print("Finished") 
